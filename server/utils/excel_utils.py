import pandas as pd
import os
from openpyxl import load_workbook
from pydantic import BaseModel

class FamilyDatabaseParams(BaseModel):
    category: str
    category_description: str
    family: str
    material: str
    material_description: str
    panel_number: str
    panel_number_description: str
    size_width: str
    size_height: str
    filepath: str

def check_type_mark(filepath, family_params: FamilyDatabaseParams):
    try:
        print (os.getcwd())
        # Load the Excel file
        df = pd.read_excel(filepath, engine='openpyxl')
        
        # Check if the required columns exist
        if 'type_mark' not in df.columns or 'filepath' not in df.columns:
            print("The required columns ('Type Mark' and 'Filepath') do not exist in the Excel file.")
            return None
        
        type_mark = f"{family_params.category}{family_params.material}-{family_params.panel_number}-{int(int(family_params.size_width)/10)}-{int(int(family_params.size_height)/10)}"

        # Check for matches in the "Type Mark" column
        matches = df[df['type_mark'] == type_mark]

        
        # Print results
        if not matches.empty:
            print(f"Found {len(matches)} match(es) for the type mark '{type_mark}':")
            print(matches)
            # Get the "Filepath" value from the first matching row
            filepath_value = matches.iloc[0]['filepath']
            return filepath_value
        else:
            print(f"No matches found for the type mark '{type_mark}'.")
            return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    

def update_excel(filepath, new_data: FamilyDatabaseParams):
    """
    Update an Excel file by adding a new row.

    :param filepath: The path to the Excel file.
    :param new_data: An instance of FamilyDatabaseParams containing the new row data.
    """
    try:
        # Load the Excel workbook
        wb = load_workbook(filepath)
        ws = wb.active  # Assuming you want to work with the active sheet

        # Convert the new data to a dictionary
        new_data_dict = new_data.model_dump()

        # Calculate the type_mark
        new_data_dict['type_mark'] = f"{new_data.category}{new_data.material}-{new_data.panel_number}-{int(int(new_data.size_width)/10)}-{int(int(new_data.size_height)/10)}"

        # Get the header row
        header = {cell.value: cell.column for cell in ws[1]}  # {column_name: column_index}
        new_row_index = ws.max_row + 1

        # Prepare the new row data
        for col_name, col_idx in header.items():
            cell = ws.cell(row=new_row_index, column=col_idx)
            if col_name in new_data_dict:
                cell.value = new_data_dict[col_name]
            else:
                cell_above = ws.cell(row=new_row_index - 1, column=col_idx)
                data_type = cell_above.data_type
                if data_type == 'f' or data_type == 'n':  # Check if the cell above contains a formula
                    cell.value = f"={cell_above.value.split('=')[-1]}"  # Copy the formula

        # Save the updated workbook
        wb.save(filepath)
        print(f"Excel file saved successfully to {filepath}.")

    except Exception as e:
        print(f"An error occurred: {e}")


def extractDatabase(filepath):
    """
    Extracts the entire database as a JSON

    :param filepath: The path to the Excel file.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(filepath, engine='openpyxl')
        
        # Convert the DataFrame to JSON
        data_dict = df.to_dict(orient='index')
        
        return data_dict
    except Exception as e:
        print(f"An error occurred: {e}")
        return None